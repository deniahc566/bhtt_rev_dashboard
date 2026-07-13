"""
Build functions ported from Scripts/export_bhtt_rev_report.py.
Writes to io.BytesIO so Streamlit can stream the download directly.
"""
import io
import pandas as pd
import openpyxl
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BLUE_DARK  = "1558D6"
_BLUE_MED   = "4A7FE8"
_BLUE_LIGHT = "C5D8FB"
_INK        = "0D1A2E"
_MUTED      = "64748B"
_WHITE      = "FFFFFF"

_BORDER_COLOR = "D0D8E4"
_THIN  = Side(style="thin",   color=_BORDER_COLOR)
_THICK = Side(style="medium", color=_BLUE_DARK)
_NO    = Side(style=None)

_NUM_FMT = "#,##0"


def _thin_border(top=False, bottom=False, left=False, right=False):
    return Border(
        top=_THIN if top else _NO,
        bottom=_THIN if bottom else _NO,
        left=_THIN if left else _NO,
        right=_THIN if right else _NO,
    )


def _thick_border(top=False, bottom=False):
    return Border(
        top=_THICK if top else _THIN,
        bottom=_THICK if bottom else _THIN,
        left=_THIN,
        right=_THIN,
    )


def _build_sheet_ban_bhtt(ws, df_partner: pd.DataFrame, df_bhs: pd.DataFrame, months: list):
    n_months   = len(months)
    total_cols = 1 + n_months + 1

    col_widths = [32] + [12] * n_months + [14]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    hdr_font  = Font(bold=True, color=_WHITE, size=10)
    hdr_fill  = PatternFill("solid", fgColor=_BLUE_DARK)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_fills = [PatternFill("solid", fgColor=_WHITE), PatternFill("solid", fgColor="F8FAFD")]
    tot_fill  = PatternFill("solid", fgColor=_BLUE_LIGHT)
    tot_font  = Font(bold=True, color=_BLUE_DARK, size=10)
    _MED = Side(style="medium", color=_BLUE_DARK)

    def write_section_title(row, title):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font      = Font(bold=True, color=_WHITE, size=11)
        cell.fill      = PatternFill("solid", fgColor=_BLUE_MED)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border    = Border(top=_MED, bottom=_MED, left=_MED, right=_MED)
        ws.row_dimensions[row].height = 24

    def write_header(row, col1_label):
        headers = [col1_label] + [f"T{m}" for m in months] + ["LŨY KẾ"]
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = _thick_border(top=True)
        ws.row_dimensions[row].height = 24

    def write_data_row(row, label, by_month_series, lk, fill):
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font(color=_INK, size=10)
        c.fill      = fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _thin_border(bottom=True, left=True)
        for ci, m in enumerate(months, start=2):
            val = float(by_month_series.get(m, 0))
            c = ws.cell(row=row, column=ci, value=val if val else None)
            c.number_format = _NUM_FMT
            c.font      = Font(color=_INK, size=10)
            c.fill      = fill
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border    = _thin_border(bottom=True)
        c = ws.cell(row=row, column=total_cols, value=lk)
        c.number_format = _NUM_FMT
        c.font      = Font(bold=True, color=_BLUE_DARK, size=10)
        c.fill      = fill
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = _thin_border(bottom=True, right=True)
        ws.row_dimensions[row].height = 20

    def write_total_row(row, col_totals, grand_total):
        c = ws.cell(row=row, column=1, value="TỔNG CỘNG")
        c.font = tot_font; c.fill = tot_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _thick_border(top=True, bottom=True)
        for ci, m in enumerate(months, start=2):
            c = ws.cell(row=row, column=ci, value=col_totals[m])
            c.number_format = _NUM_FMT
            c.font = tot_font; c.fill = tot_fill
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _thick_border(top=True, bottom=True)
        c = ws.cell(row=row, column=total_cols, value=grand_total)
        c.number_format = _NUM_FMT
        c.font = tot_font; c.fill = tot_fill
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _thick_border(top=True, bottom=True)
        ws.row_dimensions[row].height = 22

    has_bhs = not df_bhs.empty

    bhs_by_month  = df_bhs.groupby("thang")["doanh_thu"].sum() if has_bhs else pd.Series(dtype=float)
    bhs_lk        = float(df_bhs["doanh_thu"].sum()) if has_bhs else 0.0
    ptdt_by_month = df_partner.groupby("thang")["tien_thuc_thu"].sum()
    ptdt_lk       = float(df_partner["tien_thuc_thu"].sum())

    bhs_pivot = df_bhs.groupby(["kenh", "thang"], as_index=False)["doanh_thu"].sum() if has_bhs else pd.DataFrame()
    kenh_lk   = (
        df_bhs.groupby("kenh", as_index=False)["doanh_thu"]
        .sum().sort_values("doanh_thu", ascending=False)
    ) if has_bhs else pd.DataFrame()

    ptdt_pivot = df_partner.groupby(["nhom_doi_tac", "thang"], as_index=False)["tien_thuc_thu"].sum()
    nhom_lk    = (
        df_partner.groupby("nhom_doi_tac", as_index=False)["tien_thuc_thu"]
        .sum().sort_values("tien_thuc_thu", ascending=False)
    )

    r = 1

    write_section_title(r, "DOANH THU BAN BHTT"); r += 1
    write_header(r, "BAN"); r += 1

    if has_bhs:
        write_data_row(r, "Ban BHS",  bhs_by_month,  bhs_lk,  row_fills[0]); r += 1
    write_data_row(r, "Ban PTĐT", ptdt_by_month, ptdt_lk, row_fills[1 if has_bhs else 0]); r += 1

    tong_by_month = {m: float(bhs_by_month.get(m, 0)) + float(ptdt_by_month.get(m, 0)) for m in months}
    write_total_row(r, tong_by_month, bhs_lk + ptdt_lk); r += 1
    r += 2

    if has_bhs:
        write_section_title(r, "DOANH THU BAN BHS THEO KÊNH"); r += 1
        write_header(r, "KÊNH"); r += 1

        bhs_month_totals = {m: 0.0 for m in months}
        for ri_off, (_, kr) in enumerate(kenh_lk.iterrows()):
            kenh = kr["kenh"]
            lk   = float(kr["doanh_thu"])
            by_m = bhs_pivot[bhs_pivot["kenh"] == kenh].set_index("thang")["doanh_thu"]
            for m in months:
                bhs_month_totals[m] += float(by_m.get(m, 0))
            write_data_row(r, kenh, by_m, lk, row_fills[ri_off % 2]); r += 1

        write_total_row(r, bhs_month_totals, sum(bhs_month_totals.values())); r += 1
        r += 2

    write_section_title(r, "DOANH THU BAN PTĐT THEO NHÓM ĐỐI TÁC"); r += 1
    write_header(r, "NHÓM ĐỐI TÁC"); r += 1

    ptdt_month_totals = {m: 0.0 for m in months}
    for ri_off, (_, nr) in enumerate(nhom_lk.iterrows()):
        nhom = nr["nhom_doi_tac"]
        lk   = float(nr["tien_thuc_thu"])
        by_m = ptdt_pivot[ptdt_pivot["nhom_doi_tac"] == nhom].set_index("thang")["tien_thuc_thu"]
        for m in months:
            ptdt_month_totals[m] += float(by_m.get(m, 0))
        write_data_row(r, nhom, by_m, lk, row_fills[ri_off % 2]); r += 1

    write_total_row(r, ptdt_month_totals, sum(ptdt_month_totals.values()))
    ws.freeze_panes = "B1"


def _build_sheet2(ws, df: pd.DataFrame, months: list):
    n_months   = len(months)
    total_cols = 1 + n_months + 1

    hdr_font  = Font(bold=True, color=_WHITE, size=10)
    hdr_fill  = PatternFill("solid", fgColor=_BLUE_DARK)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers    = ["ĐỐI TÁC"] + [f"T{m}" for m in months] + ["LŨY KẾ"]
    col_widths = [36] + [12] * n_months + [14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _thick_border(top=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    detail_agg = df.groupby(["nhom_doi_tac", "doi_tac", "thang"], as_index=False)["tien_thuc_thu"].sum()
    nhom_agg   = df.groupby(["nhom_doi_tac", "thang"], as_index=False)["tien_thuc_thu"].sum()
    detail_lk  = (
        df.groupby(["nhom_doi_tac", "doi_tac"], as_index=False)["tien_thuc_thu"]
        .sum().rename(columns={"tien_thuc_thu": "lk"})
    )
    nhom_total = detail_lk.groupby("nhom_doi_tac")["lk"].sum().rename("nhom_lk")
    detail_lk  = detail_lk.merge(nhom_total, on="nhom_doi_tac")
    detail_lk["_nhom_khac"] = detail_lk["nhom_doi_tac"].str.upper().str.contains("KHÁC|KHAC", na=False)
    detail_lk["_is_khac"]   = detail_lk["doi_tac"].str.lower().str.contains("khác|khac", na=False)
    detail_lk = detail_lk.sort_values(
        ["_nhom_khac", "nhom_lk", "_is_khac", "lk"],
        ascending=[True, False, True, False]
    ).drop(columns=["nhom_lk", "_nhom_khac", "_is_khac"])

    _prod_df = df[df["product_group"].str.strip() != ""]
    prod_lk  = (
        _prod_df.groupby(["doi_tac", "product_group"], as_index=False)["tien_thuc_thu"]
        .sum().rename(columns={"tien_thuc_thu": "plk"})
        .sort_values(["doi_tac", "plk"], ascending=[True, False])
    )
    prod_agg = _prod_df.groupby(["doi_tac", "product_group", "thang"], as_index=False)["tien_thuc_thu"].sum()
    dt_with_products = set(prod_lk["doi_tac"].unique())

    nhom_fill = PatternFill("solid", fgColor="EEF3FF")
    nhom_font = Font(bold=True, color=_BLUE_DARK, size=10)
    dt_font   = Font(color=_INK, size=10)
    prod_fill = PatternFill("solid", fgColor="F8FAFD")
    prod_font = Font(color=_MUTED, size=9, italic=True)
    tot_fill  = PatternFill("solid", fgColor=_BLUE_LIGHT)
    tot_font  = Font(bold=True, color=_BLUE_DARK, size=10)

    current_nhom = None
    ri = 2
    col_totals   = {m: 0.0 for m in months}
    col_total_lk = 0.0

    for row in detail_lk.itertuples(index=False):
        nhom = row.nhom_doi_tac
        dt   = row.doi_tac
        lk   = row.lk
        col_total_lk += lk

        if nhom != current_nhom:
            current_nhom  = nhom
            nhom_lk_val   = detail_lk[detail_lk["nhom_doi_tac"] == nhom]["lk"].sum()
            nhom_border_top = Side(style="medium", color=_BORDER_COLOR)

            c = ws.cell(row=ri, column=1, value=nhom)
            c.font      = nhom_font
            c.fill      = nhom_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = Border(top=nhom_border_top, bottom=_THIN, left=_THIN, right=_NO)

            for ci, m in enumerate(months, start=2):
                nhom_m = float(
                    nhom_agg[(nhom_agg["nhom_doi_tac"] == nhom) & (nhom_agg["thang"] == m)]["tien_thuc_thu"].sum()
                )
                c2 = ws.cell(row=ri, column=ci, value=nhom_m if nhom_m else None)
                c2.number_format = _NUM_FMT
                c2.font      = nhom_font
                c2.fill      = nhom_fill
                c2.alignment = Alignment(horizontal="right", vertical="center")
                c2.border    = Border(top=nhom_border_top, bottom=_THIN, left=_NO, right=_NO)

            c_lk = ws.cell(row=ri, column=total_cols, value=nhom_lk_val)
            c_lk.number_format = _NUM_FMT
            c_lk.font      = nhom_font
            c_lk.fill      = nhom_fill
            c_lk.alignment = Alignment(horizontal="right", vertical="center")
            c_lk.border    = Border(top=nhom_border_top, bottom=_THIN, left=_NO, right=_THIN)
            ws.row_dimensions[ri].height = 20
            ri += 1

        has_prod = dt in dt_with_products
        name_val = f"  ▸ {dt}" if has_prod else f"    {dt}"
        c = ws.cell(row=ri, column=1, value=name_val)
        c.font      = dt_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _thin_border(bottom=True, left=True)

        for ci, m in enumerate(months, start=2):
            m_val = float(
                detail_agg[(detail_agg["doi_tac"] == dt) & (detail_agg["thang"] == m)]["tien_thuc_thu"].sum()
            )
            col_totals[m] += m_val
            c = ws.cell(row=ri, column=ci, value=m_val if m_val else None)
            c.number_format = _NUM_FMT
            c.font      = dt_font
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border    = _thin_border(bottom=True)

        c = ws.cell(row=ri, column=total_cols, value=lk)
        c.number_format = _NUM_FMT
        c.font      = dt_font
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = _thin_border(bottom=True, right=True)
        ws.row_dimensions[ri].height = 18
        ri += 1

        if has_prod:
            prods = prod_lk[prod_lk["doi_tac"] == dt].sort_values("plk", ascending=False)
            for _, pr in prods.iterrows():
                pg  = pr["product_group"]
                plk = pr["plk"]

                c = ws.cell(row=ri, column=1, value=f"        {pg}")
                c.font      = prod_font
                c.fill      = prod_fill
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border    = _thin_border(bottom=True, left=True)

                for ci, m in enumerate(months, start=2):
                    pv = float(
                        prod_agg[
                            (prod_agg["doi_tac"] == dt) & (prod_agg["product_group"] == pg)
                            & (prod_agg["thang"] == m)
                        ]["tien_thuc_thu"].sum()
                    )
                    c = ws.cell(row=ri, column=ci, value=pv if pv else None)
                    c.number_format = _NUM_FMT
                    c.font      = prod_font
                    c.fill      = prod_fill
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.border    = _thin_border(bottom=True)

                c = ws.cell(row=ri, column=total_cols, value=plk)
                c.number_format = _NUM_FMT
                c.font      = prod_font
                c.fill      = prod_fill
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.border    = _thin_border(bottom=True, right=True)
                ws.row_dimensions[ri].height = 16
                ri += 1

    c = ws.cell(row=ri, column=1, value="TỔNG CỘNG")
    c.font = tot_font; c.fill = tot_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _thick_border(top=True, bottom=True)

    for ci, m in enumerate(months, start=2):
        c = ws.cell(row=ri, column=ci, value=col_totals[m])
        c.number_format = _NUM_FMT
        c.font = tot_font; c.fill = tot_fill
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _thick_border(top=True, bottom=True)

    c = ws.cell(row=ri, column=total_cols, value=col_total_lk)
    c.number_format = _NUM_FMT
    c.font = tot_font; c.fill = tot_fill
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _thick_border(top=True, bottom=True)
    ws.row_dimensions[ri].height = 22
    ws.freeze_panes = "B2"


@st.cache_data(ttl=300, show_spinner=False)
def generate_excel_bytes(
    df_partner: pd.DataFrame,
    df_bhs: pd.DataFrame,
    year: int,
    month: int | None,
) -> bytes:
    df  = df_partner[df_partner["nam"] == year].copy()
    dfb = df_bhs[df_bhs["nam"] == year].copy()

    if month is not None:
        df  = df[df["thang"] <= month].copy()
        dfb = dfb[dfb["thang"] <= month].copy()

    months = sorted(df["thang"].unique().tolist())

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Ban BHTT"
    _build_sheet_ban_bhtt(ws1, df, dfb, months)

    ws2 = wb.create_sheet("Chi tiết đối tác")
    _build_sheet2(ws2, df, months)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
